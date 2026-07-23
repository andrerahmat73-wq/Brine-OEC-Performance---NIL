"""
Convert the OEC brine performance workbook (Pivision export) into the
data/oec_data.json file the dashboard reads.

Usage:
    python convert_excel.py path/to/your_export.xlsx

Run this every time you have a new Excel export, then commit the updated
data/oec_data.json file to the repo. GitHub Pages will pick it up on the
next page load (no rebuild step needed).

Requires: openpyxl  (pip install openpyxl)
"""

import sys
import json
import datetime
import os
from openpyxl import load_workbook

SHEETS = ['OEC 1', 'OEC 2', 'OEC 3', 'OEC 4']
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'oec_data.json')


def find_col(headers, name, occurrence='first'):
    idxs = [i for i, h in enumerate(headers)
            if h is not None and str(h).strip().lower() == name.lower()]
    if not idxs:
        return None
    return idxs[0] if occurrence == 'first' else idxs[-1]


def find_any(headers, names, occurrence='first'):
    for n in names:
        c = find_col(headers, n, occurrence)
        if c is not None:
            return c
    return None


def extract_sheet(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_time = find_col(headers, 'Time')
    # 'U' / 'Ud' / 'UpHE' appears twice: first = vaporizer, second = preheater
    u_idxs = [i for i, h in enumerate(headers) if h in ('U', 'Ud', 'UpHE')]
    u_vap_col = u_idxs[0] if len(u_idxs) >= 1 else None
    u_phe_col = u_idxs[1] if len(u_idxs) >= 2 else None

    uc_vap_col = find_col(headers, 'Uc vap', 'last')
    rd_vap_col = find_col(headers, 'Rd vap', 'last')
    uc_phe_col = find_any(headers, ['Uc PreHE', 'Uc Preheater'], 'last')
    rd_phe_col = find_col(headers, 'Rd PreHe', 'last')

    # 'MTD' appears 2-3 times: 1st = vaporizer block, 2nd = preheater block,
    # last = the binned/categorized summary block (not used here)
    mtd_idxs = [i for i, h in enumerate(headers) if h is not None and str(h).strip().lower() == 'mtd']
    mtd_vap_col = mtd_idxs[0] if len(mtd_idxs) >= 1 else None
    mtd_phe_col = mtd_idxs[1] if len(mtd_idxs) >= 2 else None
    genpower_col = find_col(headers, 'Pane1-Generator Gross Power')
    ambient_col = find_col(headers, 'Pane1-Ambient Temp')
    flow_col = find_any(headers, [
        'Pane1-OEC-1 Brine Inlet Compensated Flow',
        'Pane1-OEC-2 Brine Inlet Flow Comp',
    ])

    def g(ws, r, col, nd=3):
        if col is None:
            return None
        v = ws.cell(row=r, column=col + 1).value
        if isinstance(v, (int, float)):
            return round(v, nd)
        if isinstance(v, str) and v.startswith('#'):
            return None
        if isinstance(v, str) and v.strip().lower() in ('no data', 'nodata', ''):
            return None
        return v

    rows = []
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=col_time + 1).value
        if t is None:
            continue
        date_str = t.isoformat()[:10] if isinstance(t, (datetime.datetime, datetime.date)) else str(t)[:10]
        rows.append({
            'date': date_str,
            'u_vap': g(ws, r, u_vap_col),
            'uc_vap': g(ws, r, uc_vap_col),
            'rd_vap': g(ws, r, rd_vap_col, 6),
            'mtd_vap': g(ws, r, mtd_vap_col),
            'u_preheater': g(ws, r, u_phe_col),
            'uc_preheater': g(ws, r, uc_phe_col),
            'rd_preheater': g(ws, r, rd_phe_col, 6),
            'mtd_preheater': g(ws, r, mtd_phe_col),
            'gross_power_mw': g(ws, r, genpower_col),
            'ambient_temp': g(ws, r, ambient_col),
            'brine_flow': g(ws, r, flow_col),
        })
    return rows


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    wb = load_workbook(src, data_only=True)

    result = {}
    for sn in SHEETS:
        if sn not in wb.sheetnames:
            print(f'WARNING: sheet "{sn}" not found, skipping')
            continue
        rows = extract_sheet(wb[sn])
        key = sn.replace('OEC ', 'oec')
        result[key] = rows
        print(f'{sn}: extracted {len(rows)} rows '
              f'({rows[0]["date"] if rows else "?"} to {rows[-1]["date"] if rows else "?"})')

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(result, f, indent=0)
    print(f'\nWrote {OUTPUT_PATH}')
    print('Next: git add data/oec_data.json && git commit -m "Update OEC data" && git push')


if __name__ == '__main__':
    main()
